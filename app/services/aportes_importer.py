"""
Importador de aportes monotributo SERVIRED desde archivos Excel.

Lee archivos .xlsx con la estructura de aportes mensuales por categoría
de monotributo (A a K) y carga los registros en la tabla
ServiredAportesMonotributoDB usando upsert (no duplica, actualiza
si el monto cambió).

Uso:
    # CLI
    python -m app.services.aportes_importer servired_knowledge/aportes/aportes_monotributo.xlsx

    # Programático
    from app.services.aportes_importer import importar_aportes
    from app.database.database import get_session_factory
    Session = get_session_factory()
    db = Session()
    resultado = importar_aportes("archivo.xlsx", db)
    print(resultado)
"""

from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.database.repository import AportesMonotributoRepository

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# Resultado
# ─────────────────────────────────────────────

@dataclass
class ImportResult:
    """Resultado de la importación de aportes."""
    archivo: str
    registros_creados: int = 0
    registros_actualizados: int = 0
    registros_sin_cambio: int = 0
    errores: list[str] = field(default_factory=list)
    registros_totales: int = 0

    @property
    def exitoso(self) -> bool:
        return len(self.errores) == 0

    def resumen(self) -> str:
        lineas = [
            f"Archivo: {self.archivo}",
            f"Creados: {self.registros_creados}",
            f"Actualizados: {self.registros_actualizados}",
            f"Sin cambio: {self.registros_sin_cambio}",
            f"Total procesados: {self.registros_totales}",
        ]
        if self.errores:
            lineas.append(f"Errores: {len(self.errores)}")
            for err in self.errores:
                lineas.append(f"  - {err}")
        return "\n".join(lineas)


# ─────────────────────────────────────────────
# Parseo
# ─────────────────────────────────────────────

_CATEGORIAS_VALIDAS = set("ABCDEFGHIJK")


def _normalizar_categoria(valor: Any) -> str | None:
    """
    Normaliza un valor a categoría de monotributo válida.

    Acepta: "A", "a", " Cat A ", "A ", etc.
    Returns:
        Letra mayúscula (A-K) o None si no es válida.
    """
    if valor is None:
        return None
    texto = str(valor).strip().upper()
    texto = re.sub(r'[^A-Z]', '', texto)
    if len(texto) == 1 and texto in _CATEGORIAS_VALIDAS:
        return texto
    return None


def _parsear_monto(valor: Any) -> float:
    """
    Parsea un valor a monto float.

    Acepta:
    - Números directos (int/float)
    - Strings con formato: "$42.737,30", "42737.30"
    """
    if isinstance(valor, (int, float)):
        return float(valor)

    if isinstance(valor, str):
        valor_limpio = valor.strip()
        valor_limpio = valor_limpio.replace("$", "")
        valor_limpio = valor_limpio.replace(" ", "")

        if "." in valor_limpio and "," in valor_limpio:
            valor_limpio = valor_limpio.replace(".", "").replace(",", ".")
        elif "," in valor_limpio:
            valor_limpio = valor_limpio.replace(",", ".")
        elif "." in valor_limpio:
            partes = valor_limpio.split(".")
            if len(partes) > 2:
                valor_limpio = "".join(partes)
            elif len(partes[1]) == 3:
                valor_limpio = "".join(partes)

        try:
            return float(valor_limpio)
        except ValueError:
            return 0.0

    return 0.0


def _buscar_columna_categoria(header: list[str]) -> int:
    """Busca el índice de la columna que contiene la categoría."""
    keywords = {"categoria", "categoría", "cat", "categ"}
    for i, col in enumerate(header):
        col_lower = str(col).strip().lower() if col else ""
        if col_lower in keywords:
            return i
        for kw in keywords:
            if kw in col_lower:
                return i
    return -1


def _buscar_columna_monto(header: list[str], idx_categoria: int) -> int:
    """Busca el índice de la columna que contiene el monto del aporte."""
    keywords = {"aporte", "monto", "valor", "importe", "contribucion", "contribución"}
    for i, col in enumerate(header):
        if i == idx_categoria:
            continue
        col_lower = str(col).strip().lower() if col else ""
        if col_lower in keywords:
            return i
        for kw in keywords:
            if kw in col_lower:
                return i
    # Si no encontró por keyword, usar la siguiente columna al lado
    if idx_categoria + 1 < len(header):
        return idx_categoria + 1
    return -1


# ─────────────────────────────────────────────
# Importación principal
# ─────────────────────────────────────────────

def importar_aportes(
    ruta_excel: str | Path,
    db: Session,
) -> ImportResult:
    """
    Importa aportes monotributo desde un archivo Excel a la DB.

    Lee el Excel, parsea categorías (A-K) y montos, y hace upsert
    en ServiredAportesMonotributoDB.

    Args:
        ruta_excel: Ruta al archivo .xlsx
        db: Sesión de SQLAlchemy

    Returns:
        ImportResult con estadísticas de la importación.
    """
    ruta = Path(ruta_excel)
    resultado = ImportResult(archivo=ruta.name)

    if not ruta.exists():
        resultado.errores.append(f"Archivo no encontrado: {ruta}")
        return resultado

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(ruta), read_only=True, data_only=True)
    except ImportError:
        resultado.errores.append(
            "Para leer archivos .xlsx, instalá openpyxl: pip install openpyxl"
        )
        return resultado

    try:
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not filas or len(filas) < 2:
        resultado.errores.append("Archivo vacío o sin datos")
        return resultado

    repo = AportesMonotributoRepository(db)
    nombre_fuente = ruta.name

    # Parsear header
    header = [str(c).strip() if c else "" for c in filas[0]]
    idx_categoria = _buscar_columna_categoria(header)
    idx_monto = _buscar_columna_monto(header, idx_categoria)

    if idx_categoria == -1:
        resultado.errores.append("No se encontró columna de categoría")
        return resultado

    if idx_monto == -1:
        resultado.errores.append("No se encontró columna de monto/aporte")
        return resultado

    # Procesar filas de datos
    for fila_idx, fila in enumerate(filas[1:], start=2):
        if not fila or not any(fila):
            continue

        categoria = _normalizar_categoria(fila[idx_categoria])
        if categoria is None:
            continue

        monto = _parsear_monto(fila[idx_monto])
        if monto <= 0:
            continue

        try:
            accion, _registro = repo.upsert(
                categoria=categoria,
                monto=monto,
                fuente=nombre_fuente,
            )
            resultado.registros_totales += 1

            if accion == "created":
                resultado.registros_creados += 1
            elif accion == "updated":
                resultado.registros_actualizados += 1
            else:
                resultado.registros_sin_cambio += 1

        except Exception as exc:
            msg = f"Fila {fila_idx}, categoría '{categoria}': {exc}"
            resultado.errores.append(msg)
            logger.error("[APORTES_IMPORT] %s", msg)

    logger.info(
        "[APORTES_IMPORT] Importación completada: %s",
        resultado.resumen(),
    )
    return resultado


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def main() -> int:
    """
    CLI para importar aportes monotributo SERVIRED.

    Uso:
        python -m app.services.aportes_importer <archivo.xlsx>
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if len(sys.argv) < 2:
        print("Uso: python -m app.services.aportes_importer <archivo.xlsx>")
        print()
        print("Ejemplo:")
        print("  python -m app.services.aportes_importer servired_knowledge/aportes/aportes_monotributo.xlsx")
        print()
        print("Lee un archivo .xlsx con aportes monotributo por categoría")
        print("y los carga en la tabla servired_aportes_monotributo (upsert).")
        return 1

    ruta_excel = sys.argv[1]
    ruta = Path(ruta_excel)

    if not ruta.exists():
        print(f"Error: Archivo no encontrado: {ruta}")
        return 1

    import os
    from app.database.database import get_engine, get_session_factory, crear_tablas

    logger.info("Conectando a la base de datos...")
    database_url = os.environ.get("DATABASE_URL")
    engine = get_engine(database_url)
    crear_tablas(engine)
    SessionLocal = get_session_factory(engine)
    db = SessionLocal()

    try:
        resultado = importar_aportes(ruta, db)

        print()
        print("=" * 60)
        print("  IMPORTACIÓN DE APORTES MONOTRIBUTOS SERVIRED")
        print("=" * 60)
        print(f"  {resultado.resumen()}")
        print("=" * 60)

        return 0 if resultado.exitoso else 1

    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
