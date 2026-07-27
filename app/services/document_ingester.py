"""
Ingestor de documentos para la base de conocimiento SERVIRED.

Convierte archivos markdown, texto, PDF y XLSX en registros
de la tabla unificada ServiredKnowledgeDB.

Uso:
    # Un archivo individual
    from app.services.document_ingester import DocumentIngester
    ingester = DocumentIngester(knowledge_engine)
    ingester.ingestir_markdown("planes", "docs/planes.md")
    ingester.ingestir_texto("coberturas", "Coberturas SERVIRED", "...")

    # Carpeta completa
    ingester.ingestir_carpeta("servired_knowledge/")

    # CLI
    python -m app.services.document_ingester servired_knowledge/
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Palabras clave para detectar categoría desde el nombre del archivo
_CATEGORIAS_KEYWORDS: dict[str, list[str]] = {
    "planes": ["plan", "planes", "medimax", "gold", "medimaxgold"],
    "coberturas": ["cobertura", "coberturas", "ambulatorio", "internacion"],
    "beneficios": ["beneficio", "beneficios", "descuento", "descuentos"],
    "objeciones": ["objecion", "objeciones", "rechazo", "duda"],
    "argumentos": ["argumento", "argumentos", "pitch", "presentacion"],
    "cierres": ["cierre", "cierres", "closing", "_venta"],
    "precios": ["precio", "precios", "costo", "costos", "tarifa"],
    "informacion": ["info", "informacion", "general", "empresa", "servired"],
}


def _detectar_categoria(nombre_archivo: str) -> str:
    """Detecta la categoría a partir del nombre de archivo."""
    nombre_lower = nombre_archivo.lower().replace("-", "_").replace(" ", "_")
    for categoria, keywords in _CATEGORIAS_KEYWORDS.items():
        for kw in keywords:
            if kw in nombre_lower:
                return categoria
    return "informacion"


class DocumentIngester:
    """
    Pipeline de ingestión de documentos para KnowledgeEngine.

    Convierte archivos en registros de ServiredKnowledgeDB.
    Para archivos Excel de precios, crea registros estructurados
    en ServiredPriceDB.
    """

    def __init__(self, knowledge_engine, price_repository=None) -> None:
        """
        Args:
            knowledge_engine: KnowledgeEngine para documentos de texto.
            price_repository: PriceRepository para precios Excel (opcional).
        """
        self._engine = knowledge_engine
        self._price_repo = price_repository

    # ─────────────────────────────────────────
    # Ingesta por tipo de archivo
    # ─────────────────────────────────────────

    def ingestir_markdown(
        self,
        categoria: str,
        ruta_archivo: str | Path,
        titulo: Optional[str] = None,
        tags: str = "",
        prioridad_comercial: int = 0,
    ) -> int:
        """
        Ingesta un archivo markdown como registro de conocimiento.

        Returns:
            ID del registro creado.
        """
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        contenido = ruta.read_text(encoding="utf-8")
        if titulo is None:
            titulo = ruta.stem.replace("_", " ").replace("-", " ").title()

        logger.info("[DOC] Markdown encontrado: %s", ruta.name)
        item_id = self._engine.guardar(
            titulo=titulo,
            categoria=categoria,
            contenido=contenido,
            tags=tags,
            fuente=str(ruta),
            prioridad_comercial=prioridad_comercial,
        )
        logger.info(
            "[DOC] Guardado en DB: id=%d, titulo='%s', categoria='%s'",
            item_id, titulo, categoria,
        )
        return item_id

    def ingestir_texto(
        self,
        categoria: str,
        titulo: str,
        contenido: str,
        tags: str = "",
        fuente: str = "",
        prioridad_comercial: int = 0,
    ) -> int:
        """
        Ingesta texto plano como registro de conocimiento.

        Returns:
            ID del registro creado.
        """
        item_id = self._engine.guardar(
            titulo=titulo,
            categoria=categoria,
            contenido=contenido,
            tags=tags,
            fuente=fuente,
            prioridad_comercial=prioridad_comercial,
        )
        logger.info(
            "[DOC] Guardado en DB: id=%d, titulo='%s', categoria='%s'",
            item_id, titulo, categoria,
        )
        return item_id

    def ingestir_txt(
        self,
        categoria: str,
        ruta_archivo: str | Path,
        titulo: Optional[str] = None,
        tags: str = "",
        prioridad_comercial: int = 0,
    ) -> int:
        """
        Ingesta un archivo .txt como registro de conocimiento.

        Returns:
            ID del registro creado.
        """
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        contenido = ruta.read_text(encoding="utf-8")
        if titulo is None:
            titulo = ruta.stem.replace("_", " ").replace("-", " ").title()

        logger.info("[DOC] Texto encontrado: %s", ruta.name)
        item_id = self._engine.guardar(
            titulo=titulo,
            categoria=categoria,
            contenido=contenido,
            tags=tags,
            fuente=str(ruta),
            prioridad_comercial=prioridad_comercial,
        )
        logger.info(
            "[DOC] Guardado en DB: id=%d, titulo='%s', categoria='%s'",
            item_id, titulo, categoria,
        )
        return item_id

    def ingestir_pdf(
        self,
        categoria: str,
        ruta_archivo: str | Path,
        titulo: Optional[str] = None,
        tags: str = "",
        prioridad_comercial: int = 0,
    ) -> int:
        """
        Ingesta un PDF como registro de conocimiento.

        Requiere PyPDF2 o pdfplumber.
        """
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        if titulo is None:
            titulo = ruta.stem.replace("_", " ").replace("-", " ").title()

        logger.info("[DOC] PDF encontrado: %s", ruta.name)
        contenido = self._extraer_texto_pdf(ruta)

        item_id = self._engine.guardar(
            titulo=titulo,
            categoria=categoria,
            contenido=contenido,
            tags=tags,
            fuente=str(ruta),
            prioridad_comercial=prioridad_comercial,
        )
        logger.info(
            "[DOC] Guardado en DB: id=%d, titulo='%s', categoria='%s'",
            item_id, titulo, categoria,
        )
        return item_id

    def ingestir_xlsx(
        self,
        categoria: str,
        ruta_archivo: str | Path,
        titulo: Optional[str] = None,
        tags: str = "",
        prioridad_comercial: int = 0,
    ) -> int:
        """
        Ingesta un archivo .xlsx como registro de conocimiento.

        Lee cada fila como una línea de contenido. Si tiene columnas
        'titulo' y 'contenido', las usa directamente.

        Requiere: openpyxl
        """
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise NotImplementedError(
                "Para ingestar XLSX, instalá openpyxl: pip install openpyxl"
            )

        if titulo is None:
            titulo = ruta.stem.replace("_", " ").replace("-", " ").title()

        logger.info("[DOC] XLSX encontrado: %s", ruta.name)

        wb = load_workbook(str(ruta), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"El archivo XLSX no tiene hojas activas: {ruta}")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            raise ValueError(f"El archivo XLSX está vacío: {ruta}")

        header = [str(c).strip().lower() if c else "" for c in rows[0]]

        # Si tiene columnas titulo/contenido, usarlas directamente
        if "titulo" in header and "contenido" in header:
            idx_titulo = header.index("titulo")
            idx_contenido = header.index("contenido")
            lineas = []
            for row in rows[1:]:
                t = str(row[idx_titulo]) if row[idx_titulo] else ""
                c = str(row[idx_contenido]) if row[idx_contenido] else ""
                if t or c:
                    lineas.append(f"{t}: {c}" if t and c else (t or c))
            contenido = "\n".join(lineas)
        else:
            # Sin columnas conocidas, volcar todo como texto
            lineas = []
            for row in rows:
                vals = [str(c) if c else "" for c in row]
                if any(vals):
                    lineas.append(" | ".join(vals))
            contenido = "\n".join(lineas)

        wb.close()

        item_id = self._engine.guardar(
            titulo=titulo,
            categoria=categoria,
            contenido=contenido,
            tags=tags,
            fuente=str(ruta),
            prioridad_comercial=prioridad_comercial,
        )
        logger.info(
            "[DOC] Guardado en DB: id=%d, titulo='%s', categoria='%s', filas=%d",
            item_id, titulo, categoria, len(rows) - 1,
        )
        return item_id

    # ─────────────────────────────────────────
    # Ingesta de precios estructurados
    # ─────────────────────────────────────────

    def ingestir_xlsx_precios(
        self,
        ruta_archivo: str | Path,
        tipo_afiliacion: str,
    ) -> int:
        """
        Ingesta un archivo Excel de precios como registros estructurados.

        Crea registros en ServiredPriceDB en lugar de ServiredKnowledgeDB.
        Cada fila del Excel se convierte en un registro de precio.

        Formato esperado del Excel:
        - Hoja: nombre del tipo de afiliación (particular, monotributo, etc.)
        - Columna A: Nombre del plan
        - Columnas siguientes: precios por zona/rango de edad
          (ej: "18-30 Córdoba", "18-30 Interior", "31+ Córdoba", "31+ Interior")
          O simplificado: "Córdoba", "Interior"

        Args:
            ruta_archivo: Ruta al archivo .xlsx
            tipo_afiliacion: particular|monotributo|relacion_dependencia

        Returns:
            Cantidad de registros de precio creados.
        """
        if self._price_repo is None:
            raise ValueError(
                "Se requiere PriceRepository para ingestar precios. "
                "Inicializá DocumentIngester con price_repository."
            )

        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise NotImplementedError(
                "Para ingestar XLSX, instalá openpyxl: pip install openpyxl"
            )

        logger.info(
            "[DOC] XLSX precios encontrado: %s (tipo=%s)",
            ruta.name, tipo_afiliacion,
        )

        wb = load_workbook(str(ruta), read_only=True, data_only=True)

        precios_creados = 0

        # Procesar cada hoja del libro
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Detectar columnas del header
            header = [str(c).strip() if c else "" for c in rows[0]]

            # Buscar columna de plan
            idx_plan = self._buscar_columna_plan(header)
            if idx_plan == -1:
                logger.warning(
                    "[DOC] No se encontró columna de plan en hoja '%s'",
                    ws.title,
                )
                continue

            # Mapear columnas de precios (zonas/rangos de edad)
            columnas_precio = self._mapear_columnas_precio(header, idx_plan)

            if not columnas_precio:
                logger.warning(
                    "[DOC] No se encontraron columnas de precio en hoja '%s'",
                    ws.title,
                )
                continue

            # Procesar filas de datos
            for row in rows[1:]:
                if not row or not any(row):
                    continue

                plan_nombre = str(row[idx_plan]).strip() if row[idx_plan] else ""
                if not plan_nombre or plan_nombre.lower() in ("plan", "planes", ""):
                    continue

                plan_normalizado = self._normalizar_nombre_plan(plan_nombre)

                for col_info in columnas_precio:
                    valor = row[col_info["idx"]]
                    if valor is None:
                        continue

                    try:
                        precio = self._parsear_precio(valor)
                    except (ValueError, TypeError):
                        continue

                    if precio <= 0:
                        continue

                    # Crear registro de precio
                    self._price_repo.crear(
                        tipo_afiliacion=tipo_afiliacion,
                        plan=plan_normalizado,
                        zona=col_info["zona"],
                        precio=precio,
                        edad_desde=col_info.get("edad_desde", 0),
                        edad_hasta=col_info.get("edad_hasta", 99),
                        fuente=ruta.name,
                    )
                    precios_creados += 1

        wb.close()

        logger.info(
            "[DOC] Precios creados: %d registros (tipo=%s, archivo=%s)",
            precios_creados, tipo_afiliacion, ruta.name,
        )
        return precios_creados

    def _buscar_columna_plan(self, header: list[str]) -> int:
        """Busca el índice de la columna que contiene el nombre del plan."""
        keywords_plan = ["plan", "planes", "nombre", "producto"]
        for i, col in enumerate(header):
            col_lower = col.lower().strip()
            if col_lower in keywords_plan:
                return i
            for kw in keywords_plan:
                if kw in col_lower:
                    return i
        return -1

    def _mapear_columnas_precio(
        self, header: list[str], idx_plan: int
    ) -> list[dict]:
        """
        Mapea las columnas de precio del header.

        Detecta automáticamente el formato:
        - Simplificado: "Córdoba", "Interior"
        - Con rango: "18-30 Córdoba", "31+ Interior"
        - Completo: "Particulares 18-30 Córdoba"

        Returns:
            Lista de dicts con zona, edad_desde, edad_hasta, idx.
        """
        import re

        columnas = []

        for i, col in enumerate(header):
            if i == idx_plan:
                continue

            col_lower = col.lower().strip()
            if not col_lower:
                continue

            info = {"idx": i, "zona": "", "edad_desde": 0, "edad_hasta": 99}

            # Detectar zona
            if "córdoba" in col_lower or "cordoba" in col_lower:
                info["zona"] = "cordoba"
            elif "interior" in col_lower:
                info["zona"] = "interior"
            else:
                # Si no detecta zona, saltar esta columna
                continue

            # Detectar rango de edad
            patron_edad = re.search(r'(\d+)[\s]*[-+][\s]*(\d+|\+)?', col_lower)
            if patron_edad:
                edad_min = int(patron_edad.group(1))
                if "+" in col_lower:
                    info["edad_desde"] = edad_min
                    info["edad_hasta"] = 99
                elif patron_edad.group(2):
                    edad_max = int(patron_edad.group(2))
                    info["edad_desde"] = edad_min
                    info["edad_hasta"] = edad_max

            columnas.append(info)

        return columnas

    def _normalizar_nombre_plan(self, nombre: str) -> str:
        """
        Normaliza el nombre del plan a formato estándar.

        Ejemplos:
            - "Medimax CO" -> "medimax_co"
            - "Medimax Gold" -> "medimax_gold"
            - "Gold" -> "gold"
            - "Plan Joven" -> "plan_joven"
        """
        nombre_lower = nombre.lower().strip()
        nombre_limpio = nombre_lower.replace(" ", "_")

        # Mapeo de nombres comunes
        mapeo = {
            "medimax_co": "medimax_co",
            "medimaxco": "medimax_co",
            "medimax_co_": "medimax_co",
            "medimax": "medimax",
            "medimax_gold": "medimax_gold",
            "medimaxgold": "medimax_gold",
            "gold": "gold",
            "plan_joven": "plan_joven",
            "joven": "plan_joven",
        }

        return mapeo.get(nombre_limpio, nombre_limpio)

    def _parsear_precio(self, valor) -> float:
        """
        Parsea un valor a precio float.

        Acepta:
        - Números directos (int/float)
        - Strings con formato: "$15.000", "15.000", "15,5"
        """
        if isinstance(valor, (int, float)):
            return float(valor)

        if isinstance(valor, str):
            # Limpiar formato argentino
            valor_limpio = valor.strip()
            valor_limpio = valor_limpio.replace("$", "")
            valor_limpio = valor_limpio.replace(" ", "")

            # Formato argentino: 15.000 -> 15000, 15,5 -> 15.5
            if "." in valor_limpio and "," in valor_limpio:
                # 1.234,56 -> 1234.56
                valor_limpio = valor_limpio.replace(".", "").replace(",", ".")
            elif "," in valor_limpio:
                # 15,5 -> 15.5
                valor_limpio = valor_limpio.replace(",", ".")
            elif "." in valor_limpio:
                # 15.000 -> 15000 (asumimos que es separador de miles)
                partes = valor_limpio.split(".")
                if len(partes) > 2:
                    # 15.000.000 -> 15000000
                    valor_limpio = "".join(partes)
                # Si tiene exactamente 1 punto y 3 dígitos después, es separador de miles
                elif len(partes[1]) == 3:
                    valor_limpio = "".join(partes)

            return float(valor_limpio)

        return 0.0

    # ─────────────────────────────────────────
    # Ingesta masiva por carpeta
    # ─────────────────────────────────────────

    def ingestir_carpeta(
        self,
        carpeta: str | Path,
        prioridad_comercial: int = 0,
    ) -> dict:
        """
        Ingesta todos los archivos soportados de una carpeta.

        Estructura esperada:
            servired_knowledge/
              planes/
                plan_medimax.md
              coberturas/
                coberturas.pdf
              beneficios/
                descuentos.xlsx
            o bien:
            servired_knowledge/
              planes_SERVIRED.txt

        La categoría se detecta por:
        1. Nombre de subcarpeta si el archivo está en subcarpeta
        2. Nombre del archivo (keywords)

        Returns:
            Dict con estadísticas: {archivos_ok, archivos_err, ids}
        """
        carpeta_path = Path(carpeta)
        if not carpeta_path.is_dir():
            raise NotADirectoryError(f"Carpeta no encontrada: {carpeta_path}")

        stats: dict = {"archivos_ok": 0, "archivos_err": 0, "ids": []}
        extensiones = {".md", ".txt", ".pdf", ".xlsx"}

        logger.info("[DOC] === Iniciando ingestión de carpeta: %s ===", carpeta_path)

        archivos = sorted(carpeta_path.rglob("*"))
        for archivo in archivos:
            if not archivo.is_file():
                continue
            if archivo.suffix.lower() not in extensiones:
                logger.debug("[DOC] Saltando archivo no soportado: %s", archivo.name)
                continue

            # Detectar categoría: subcarpeta o nombre de archivo
            if archivo.parent != carpeta_path:
                categoria = archivo.parent.name.lower().replace(" ", "_")
            else:
                categoria = _detectar_categoria(archivo.stem)

            try:
                if archivo.suffix.lower() == ".md":
                    item_id = self.ingestir_markdown(
                        categoria, archivo,
                        prioridad_comercial=prioridad_comercial,
                    )
                elif archivo.suffix.lower() == ".txt":
                    item_id = self.ingestir_txt(
                        categoria, archivo,
                        prioridad_comercial=prioridad_comercial,
                    )
                elif archivo.suffix.lower() == ".pdf":
                    item_id = self.ingestir_pdf(
                        categoria, archivo,
                        prioridad_comercial=prioridad_comercial,
                    )
                elif archivo.suffix.lower() == ".xlsx":
                    item_id = self.ingestir_xlsx(
                        categoria, archivo,
                        prioridad_comercial=prioridad_comercial,
                    )
                else:
                    continue

                stats["archivos_ok"] += 1
                stats["ids"].append(item_id)

            except Exception as exc:
                stats["archivos_err"] += 1
                logger.error(
                    "[DOC] Error procesando %s: %s", archivo.name, exc,
                )

        logger.info(
            "[DOC] === Ingestión completada: %d OK, %d errores ===",
            stats["archivos_ok"],
            stats["archivos_err"],
        )
        return stats

    # ─────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────

    def _extraer_texto_pdf(self, ruta: Path) -> str:
        """Extrae texto de un PDF. Intenta PyPDF2, luego pdfplumber."""
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(ruta))
            partes = []
            for page in reader.pages:
                texto = page.extract_text()
                if texto:
                    partes.append(texto)
            return "\n\n".join(partes)
        except ImportError:
            pass

        try:
            import pdfplumber
            with pdfplumber.open(str(ruta)) as pdf:
                partes = []
                for page in pdf.pages:
                    texto = page.extract_text()
                    if texto:
                        partes.append(texto)
            return "\n\n".join(partes)
        except ImportError:
            raise NotImplementedError(
                "Para ingestar PDFs, instalá PyPDF2 o pdfplumber"
            )
