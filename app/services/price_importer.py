"""
Importador de precios SERVIRED desde archivos Excel.

Lee archivos .xls o .xlsx con la estructura de precios de SERVIRED,
detecta automáticamente las hojas y columnas, y carga los registros
en la tabla ServiredPriceDB usando upsert (no duplica, actualiza
si el precio cambió).

Uso:
    # CLI
    python -m app.services.price_importer servired_knowledge/precios/archivo.xlsx

    # Programático
    from app.services.price_importer import importar_precios
    from app.database.database import get_session_factory
    Session = get_session_factory()
    db = Session()
    resultado = importar_precios("archivo.xlsx", db)
    print(resultado)
"""

from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.database.repository import PriceRepository

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────

# Mapeo de nombres de hoja -> tipo_afiliacion
_HOJA_A_TIPO: dict[str, str] = {
    "particulares": "particular",
    "particular": "particular",
    "monotributos": "monotributo",
    "monotributo": "monotributo",
    "relacion de dependencia": "relacion_dependencia",
    "relación de dependencia": "relacion_dependencia",
    "relacion_dependencia": "relacion_dependencia",
    "dependencia": "relacion_dependencia",
}

# Columnas que pueden contener el nombre del plan
_KEYWORDS_PLAN = {"plan", "planes", "nombre", "producto", "servicio"}

# Mapeo de nombres de plan normalizados
_PLAN_MAP: dict[str, str] = {
    "medimax_co": "medimax_co",
    "medimaxco": "medimax_co",
    "medimax_co_": "medimax_co",
    "medimax": "medimax",
    "medimax_gold": "medimax_gold",
    "medimaxgold": "medimax_gold",
    "gold": "gold",
    "plan_joven": "plan_joven",
    "joven": "plan_joven",
}


# ─────────────────────────────────────────────
# Resultado
# ─────────────────────────────────────────────

@dataclass
class ImportResult:
    """Resultado de la importación de precios."""
    archivo: str
    hojas_procesadas: int = 0
    precios_creados: int = 0
    precios_actualizados: int = 0
    precios_sin_cambio: int = 0
    errores: list[str] = field(default_factory=list)
    precios_totales: int = 0

    @property
    def exitoso(self) -> bool:
        return len(self.errores) == 0

    def resumen(self) -> str:
        lineas = [
            f"Archivo: {self.archivo}",
            f"Hojas procesadas: {self.hojas_procesadas}",
            f"Precios creados: {self.precios_creados}",
            f"Precios actualizados: {self.precios_actualizados}",
            f"Precios sin cambio: {self.precios_sin_cambio}",
            f"Total procesados: {self.precios_totales}",
        ]
        if self.errores:
            lineas.append(f"Errores: {len(self.errores)}")
            for err in self.errores:
                lineas.append(f"  - {err}")
        return "\n".join(lineas)


# ─────────────────────────────────────────────
# Lectura de Excel (.xls y .xlsx)
# ─────────────────────────────────────────────

def _abrir_workbook(ruta: Path) -> tuple[Any, str]:
    """
    Abre un archivo Excel (.xls o .xlsx) y retorna (workbook, formato).

    El formato es "xls" o "xlsx" para que el caller sepa cómo procesar.
    """
    suffix = ruta.suffix.lower()

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError:
            raise NotImplementedError(
                "Para leer archivos .xls, instalá xlrd: pip install xlrd"
            )
        wb = xlrd.open_workbook(str(ruta))
        return wb, "xls"

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise NotImplementedError(
                "Para leer archivos .xlsx, instalá openpyxl: pip install openpyxl"
            )
        wb = load_workbook(str(ruta), read_only=True, data_only=True)
        return wb, "xlsx"

    raise ValueError(
        f"Formato no soportado: {suffix}. "
        f"Usá .xls o .xlsx."
    )


def _leer_hojas(wb: Any, formato: str) -> dict[str, list[list[Any]]]:
    """
    Lee todas las hojas del workbook y retorna un dict
    {nombre_hoja: [[row1], [row2], ...]}.
    """
    hojas: dict[str, list[list[Any]]] = {}

    if formato == "xls":
        for nombre in wb.sheet_names():
            hoja = wb.sheet_by_name(nombre)
            filas = []
            for i in range(hoja.nrows):
                filas.append(hoja.row_values(i))
            hojas[nombre] = filas
    else:
        for ws in wb.worksheets:
            filas = list(ws.iter_rows(values_only=True))
            hojas[ws.title] = filas

    return hojas


def _cerrar_workbook(wb: Any, formato: str) -> None:
    """Cierra el workbook."""
    if formato == "xlsx":
        wb.close()


# ─────────────────────────────────────────────
# Parseo de headers y precios
# ─────────────────────────────────────────────

def _buscar_columna_plan(header: list[str]) -> int:
    """Busca el índice de la columna que contiene el nombre del plan."""
    for i, col in enumerate(header):
        col_lower = str(col).strip().lower() if col else ""
        if col_lower in _KEYWORDS_PLAN:
            return i
        for kw in _KEYWORDS_PLAN:
            if kw in col_lower:
                return i
    return -1


def _mapear_columnas_precio(
    header: list[str], idx_plan: int
) -> list[dict]:
    """
    Mapea las columnas de precio del header.

    Detecta automáticamente zona (cordoba/interior) y rango de edad.

    Returns:
        Lista de dicts con zona, edad_desde, edad_hasta, idx.
    """
    columnas = []

    for i, col in enumerate(header):
        if i == idx_plan:
            continue

        col_str = str(col).strip().lower() if col else ""
        if not col_str:
            continue

        info: dict[str, Any] = {
            "idx": i,
            "zona": "",
            "edad_desde": 0,
            "edad_hasta": 99,
        }

        # Detectar zona
        if "cordoba" in col_str or "córdoba" in col_str:
            info["zona"] = "cordoba"
        elif "interior" in col_str:
            info["zona"] = "interior"
        else:
            continue

        # Detectar rango de edad
        patron = re.search(r'(\d+)[\s]*[-+][\s]*(\d+|\+)?', col_str)
        if patron:
            edad_min = int(patron.group(1))
            if "+" in col_str:
                info["edad_desde"] = edad_min
                info["edad_hasta"] = 99
            elif patron.group(2):
                info["edad_desde"] = edad_min
                info["edad_hasta"] = int(patron.group(2))

        columnas.append(info)

    return columnas


def _normalizar_plan(nombre: str) -> str:
    """Normaliza el nombre del plan a formato estándar."""
    nombre_lower = nombre.lower().strip()
    nombre_limpio = nombre_lower.replace(" ", "_")
    return _PLAN_MAP.get(nombre_limpio, nombre_limpio)


def _parsear_precio(valor: Any) -> float:
    """
    Parsea un valor a precio float.

    Acepta:
    - Números directos (int/float)
    - Strings con formato: "$15.000", "15.000", "15,5"
    """
    if isinstance(valor, (int, float)):
        return float(valor)

    if isinstance(valor, str):
        valor_limpio = valor.strip()
        valor_limpio = valor_limpio.replace("$", "")
        valor_limpio = valor_limpio.replace(" ", "")

        if "." in valor_limpio and "," in valor_limpio:
            valor_limpio = valor_limpio.replace(".", "").replace(",", ".")
        elif "," in valor_limpio:
            valor_limpio = valor_limpio.replace(",", ".")
        elif "." in valor_limpio:
            partes = valor_limpio.split(".")
            if len(partes) > 2:
                valor_limpio = "".join(partes)
            elif len(partes[1]) == 3:
                valor_limpio = "".join(partes)

        try:
            return float(valor_limpio)
        except ValueError:
            return 0.0

    return 0.0


def _detectar_tipo_hoja(nombre_hoja: str) -> str | None:
    """Detecta el tipo de afiliación a partir del nombre de la hoja."""
    nombre_lower = nombre_hoja.lower().strip()
    for clave, tipo in _HOJA_A_TIPO.items():
        if clave in nombre_lower:
            return tipo
    return None


# ─────────────────────────────────────────────
# Importación principal
# ─────────────────────────────────────────────

def importar_precios(
    ruta_excel: str | Path,
    db: Session,
    tipos_por_hoja: dict[str, str] | None = None,
) -> ImportResult:
    """
    Importa precios desde un archivo Excel a la DB.

    Lee cada hoja, detecta el tipo de afiliación (o usa el mapeo
    provisto), parsea headers de zona/edad, y hace upsert de cada
    precio en ServiredPriceDB.

    Args:
        ruta_excel: Ruta al archivo .xls o .xlsx
        db: Sesión de SQLAlchemy
        tipos_por_hoja: Mapping opcional {nombre_hoja: tipo_afiliacion}.
                        Si no se provee, se auto-detecta del nombre.

    Returns:
        ImportResult con estadísticas de la importación.
    """
    ruta = Path(ruta_excel)
    resultado = ImportResult(archivo=ruta.name)

    if not ruta.exists():
        resultado.errores.append(f"Archivo no encontrado: {ruta}")
        return resultado

    try:
        wb, formato = _abrir_workbook(ruta)
    except NotImplementedError as exc:
        resultado.errores.append(str(exc))
        return resultado

    try:
        hojas = _leer_hojas(wb, formato)
    finally:
        _cerrar_workbook(wb, formato)

    repo = PriceRepository(db)
    nombre_fuente = ruta.name

    for nombre_hoja, filas in hojas.items():
        if not filas or len(filas) < 2:
            logger.debug("[IMPORT] Hoja '%s' vacía o sin datos, saltando", nombre_hoja)
            continue

        # Detectar tipo de afiliación
        if tipos_por_hoja and nombre_hoja in tipos_por_hoja:
            tipo_afiliacion = tipos_por_hoja[nombre_hoja]
        else:
            tipo_afiliacion = _detectar_tipo_hoja(nombre_hoja)

        if tipo_afiliacion is None:
            logger.warning(
                "[IMPORT] No se pudo detectar tipo de afiliación "
                "para hoja '%s', saltando",
                nombre_hoja,
            )
            continue

        # Parsear header
        header = [str(c).strip() if c else "" for c in filas[0]]
        idx_plan = _buscar_columna_plan(header)

        if idx_plan == -1:
            logger.warning(
                "[IMPORT] No se encontró columna de plan en hoja '%s'",
                nombre_hoja,
            )
            continue

        columnas_precio = _mapear_columnas_precio(header, idx_plan)
        if not columnas_precio:
            logger.warning(
                "[IMPORT] No se encontraron columnas de precio en hoja '%s'",
                nombre_hoja,
            )
            continue

        resultado.hojas_procesadas += 1

        # Procesar filas de datos
        for fila_idx, fila in enumerate(filas[1:], start=2):
            if not fila or not any(fila):
                continue

            plan_nombre = str(fila[idx_plan]).strip() if fila[idx_plan] else ""
            if not plan_nombre or plan_nombre.lower() in ("plan", "planes", ""):
                continue

            plan_normalizado = _normalizar_plan(plan_nombre)

            for col_info in columnas_precio:
                valor = fila[col_info["idx"]]
                if valor is None or valor == "":
                    continue

                precio = _parsear_precio(valor)
                if precio <= 0:
                    continue

                try:
                    accion, _registro = repo.upsert(
                        tipo_afiliacion=tipo_afiliacion,
                        plan=plan_normalizado,
                        zona=col_info["zona"],
                        precio=precio,
                        edad_desde=col_info.get("edad_desde", 0),
                        edad_hasta=col_info.get("edad_hasta", 99),
                        fuente=nombre_fuente,
                    )
                    resultado.precios_totales += 1

                    if accion == "created":
                        resultado.precios_creados += 1
                    elif accion == "updated":
                        resultado.precios_actualizados += 1
                    else:
                        resultado.precios_sin_cambio += 1

                except Exception as exc:
                    msg = (
                        f"Fila {fila_idx}, hoja '{nombre_hoja}', "
                        f"plan '{plan_normalizado}': {exc}"
                    )
                    resultado.errores.append(msg)
                    logger.error("[IMPORT] %s", msg)

        logger.info(
            "[IMPORT] Hoja '%s' (%s): procesada",
            nombre_hoja, tipo_afiliacion,
        )

    logger.info(
        "[IMPORT] Importación completada: %s",
        resultado.resumen(),
    )
    return resultado


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def main() -> int:
    """
    CLI para importar precios SERVIRED.

    Uso:
        python -m app.services.price_importer <archivo.xlsx>
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if len(sys.argv) < 2:
        print("Uso: python -m app.services.price_importer <archivo.xlsx>")
        print()
        print("Ejemplo:")
        print("  python -m app.services.price_importer servired_knowledge/precios/precios.xlsx")
        print()
        print("Lee archivos .xls o .xlsx con precios SERVIRED y los")
        print("carga en la tabla servired_prices (upsert, no duplica).")
        return 1

    ruta_excel = sys.argv[1]
    ruta = Path(ruta_excel)

    if not ruta.exists():
        print(f"Error: Archivo no encontrado: {ruta}")
        return 1

    import os
    from app.database.database import get_engine, get_session_factory, crear_tablas

    logger.info("Conectando a la base de datos...")
    database_url = os.environ.get("DATABASE_URL")
    engine = get_engine(database_url)
    crear_tablas(engine)
    SessionLocal = get_session_factory(engine)
    db = SessionLocal()

    try:
        resultado = importar_precios(ruta, db)

        print()
        print("=" * 60)
        print("  IMPORTACIÓN DE PRECIOS SERVIRED")
        print("=" * 60)
        print(f"  {resultado.resumen()}")
        print("=" * 60)

        return 0 if resultado.exitoso else 1

    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
