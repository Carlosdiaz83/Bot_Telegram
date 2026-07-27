"""
Tests Sprint 17 — Precios estructurados SERVIRED.

Cubre:
    - ServiredPriceDB: modelo de precios
    - PriceRepository: CRUD y búsquedas
    - DocumentIngester.ingestir_xlsx_precios(): parseo de Excel
    - ServiredCalculator con PriceRepository
    - Datos ficticios de Excel para testing
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database.models import Base, ServiredPriceDB
from app.database.repository import PriceRepository
from app.models.lead import (
    EstadoComercial,
    Lead,
    TipoAfiliacion,
)


# ─────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────

@pytest.fixture()
def db_session():
    """Sesión de test con DB SQLite en memoria."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()
    engine.dispose()


@pytest.fixture()
def price_repo(db_session):
    """PriceRepository con sesión de test."""
    return PriceRepository(db_session)


# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def _crear_excel_precios(
    ruta: str,
    tipo_afiliacion: str = "particular",
    sheets: dict | None = None,
):
    """
    Crea un archivo Excel ficticio de precios.

    Args:
        ruta: Ruta del archivo a crear.
        tipo_afiliacion: Tipo de afiliación para la hoja.
        sheets: Dict con nombre_hoja -> lista de filas.
                Si es None, crea formato estándar.
    """
    wb = Workbook()

    if sheets is None:
        sheets = {
            tipo_afiliacion: [
                ["Plan", "18-30 Córdoba", "18-30 Interior", "31-50 Córdoba", "31-50 Interior"],
                ["Medimax CO", 12000, 10500, 15000, 13000],
                ["Medimax", 18000, 16000, 22000, 19500],
                ["Medimax Gold", 28000, 25000, 35000, 31000],
                ["Gold", 35000, 31000, 42000, 37000],
                ["Plan Joven", 8000, 7000, 0, 0],
            ]
        }

    for i, (sheet_name, rows) in enumerate(sheets.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row in rows:
            ws.append(row)

    wb.save(ruta)
    wb.close()


def _crear_excel_formato_simple(ruta: str):
    """Crea un Excel con formato simplificado (solo Córdoba/Interior)."""
    sheets = {
        "Particulares": [
            ["Plan", "Córdoba", "Interior"],
            ["Medimax CO", 12000, 10500],
            ["Medimax", 18000, 16000],
            ["Medimax Gold", 28000, 25000],
        ]
    }
    _crear_excel_precios(ruta, "particular", sheets)


def _crear_excel_multi_hojas(ruta: str):
    """Crea un Excel con 3 hojas (una por tipo de afiliación)."""
    sheets = {
        "Particulares": [
            ["Plan", "18-30 Córdoba", "18-30 Interior", "31+ Córdoba", "31+ Interior"],
            ["Medimax CO", 12000, 10500, 15000, 13000],
            ["Medimax", 18000, 16000, 22000, 19500],
        ],
        "Monotributo": [
            ["Plan", "18-30 Córdoba", "18-30 Interior", "31+ Córdoba", "31+ Interior"],
            ["Medimax CO", 11000, 9500, 14000, 12000],
            ["Medimax", 16500, 14500, 20000, 17500],
        ],
        "Relación de dependencia": [
            ["Plan", "18-30 Córdoba", "18-30 Interior", "31+ Córdoba", "31+ Interior"],
            ["Medimax CO", 10000, 8500, 13000, 11000],
            ["Medimax", 15000, 13000, 18500, 16000],
        ],
    }
    _crear_excel_precios(ruta, "particular", sheets)


def _crear_lead_mock(
    nombre: str = "Carlos",
    edad: int = 35,
    tipo_afiliacion: TipoAfiliacion = TipoAfiliacion.PARTICULAR,
    conyuge: bool = False,
    hijos: bool = False,
    cantidad_hijos: int = 0,
) -> Lead:
    """Crea un Lead mock para tests."""
    lead = Lead(
        lead_id="12345",
        nombre=nombre,
        edad=edad,
        estado_comercial=EstadoComercial.NUEVO,
        tipo_afiliacion=tipo_afiliacion,
    )
    if conyuge or hijos:
        lead.actualizar_grupo_familiar(
            conyuge=conyuge,
            hijos=hijos,
            cantidad_hijos=cantidad_hijos,
        )
    return lead


# ─────────────────────────────────────────
# Tests: ServiredPriceDB
# ─────────────────────────────────────────

class TestServiredPriceDB:
    """Tests del modelo ServiredPriceDB."""

    def test_crear_precio_basico(self):
        """Un precio se crea con valores por defecto."""
        precio = ServiredPriceDB(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
            edad_desde=0,
            edad_hasta=99,
            activo=True,
        )
        assert precio.tipo_afiliacion == "particular"
        assert precio.plan == "medimax"
        assert precio.zona == "cordoba"
        assert precio.precio == 18000.0
        assert precio.edad_desde == 0
        assert precio.edad_hasta == 99
        assert precio.activo is True

    def test_crear_precio_con_rango_edad(self):
        """Un precio con rango de edad específico."""
        precio = ServiredPriceDB(
            tipo_afiliacion="particular",
            plan="plan_joven",
            zona="cordoba",
            precio=8000.0,
            edad_desde=18,
            edad_hasta=30,
        )
        assert precio.edad_desde == 18
        assert precio.edad_hasta == 30

    def test_repr_precio(self):
        """El repr contiene la información clave."""
        precio = ServiredPriceDB(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )
        repr_str = repr(precio)
        assert "ServiredPriceDB" in repr_str
        assert "particular" in repr_str
        assert "medimax" in repr_str
        assert "cordoba" in repr_str


# ─────────────────────────────────────────
# Tests: PriceRepository
# ─────────────────────────────────────────

class TestPriceRepository:
    """Tests del repositorio de precios."""

    def test_crear_precio(self, db_session):
        """Crear un registro de precio."""
        repo = PriceRepository(db_session)
        precio = repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )
        assert precio is not None
        assert precio.tipo_afiliacion == "particular"
        assert precio.plan == "medimax"
        assert precio.zona == "cordoba"
        assert precio.precio == 18000.0

    def test_bulk_crear(self, db_session):
        """Crear múltiples precios en lote."""
        repo = PriceRepository(db_session)
        precios = [
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "cordoba", "precio": 18000.0},
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "interior", "precio": 16000.0},
            {"tipo_afiliacion": "particular", "plan": "gold", "zona": "cordoba", "precio": 35000.0},
        ]
        cantidad = repo.bulk_crear(precios)
        assert cantidad == 3

    def test_bulk_crear_vacio(self, db_session):
        """Bulk crear con lista vacía retorna 0."""
        repo = PriceRepository(db_session)
        cantidad = repo.bulk_crear([])
        assert cantidad == 0

    def test_buscar_precio_exacto(self, db_session):
        """Buscar precio exacto por tipo, plan, zona."""
        repo = PriceRepository(db_session)
        repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )
        precio = repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
        )
        assert precio is not None
        assert precio.precio == 18000.0

    def test_buscar_precio_con_edad(self, db_session):
        """Buscar precio con restricción de edad."""
        repo = PriceRepository(db_session)
        repo.crear(
            tipo_afiliacion="particular",
            plan="plan_joven",
            zona="cordoba",
            precio=8000.0,
            edad_desde=18,
            edad_hasta=30,
        )

        # Edad dentro del rango
        precio = repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="plan_joven",
            zona="cordoba",
            edad=25,
        )
        assert precio is not None
        assert precio.precio == 8000.0

        # Edad fuera del rango
        precio = repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="plan_joven",
            zona="cordoba",
            edad=35,
        )
        assert precio is None

    def test_buscar_precio_no_existe(self, db_session):
        """Buscar precio que no existe retorna None."""
        repo = PriceRepository(db_session)
        precio = repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="no_existe",
            zona="cordoba",
        )
        assert precio is None

    def test_buscar_todos(self, db_session):
        """Buscar todos los precios."""
        repo = PriceRepository(db_session)
        repo.bulk_crear([
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "cordoba", "precio": 18000.0},
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "interior", "precio": 16000.0},
            {"tipo_afiliacion": "monotributo", "plan": "medimax", "zona": "cordoba", "precio": 16500.0},
        ])

        todos = repo.buscar_todos()
        assert len(todos) == 3

        solo_particular = repo.buscar_todos(tipo_afiliacion="particular")
        assert len(solo_particular) == 2

        solo_medimax = repo.buscar_todos(plan="medimax")
        assert len(solo_medimax) == 3

        solo_cordoba = repo.buscar_todos(zona="cordoba")
        assert len(solo_cordoba) == 2

    def test_eliminar_por_fuente(self, db_session):
        """Eliminar precios por fuente específica."""
        repo = PriceRepository(db_session)
        repo.bulk_crear([
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "cordoba", "precio": 18000.0, "fuente": "test.xlsx"},
            {"tipo_afiliacion": "particular", "plan": "medimax", "zona": "interior", "precio": 16000.0, "fuente": "test.xlsx"},
            {"tipo_afiliacion": "particular", "plan": "gold", "zona": "cordoba", "precio": 35000.0, "fuente": "otro.xlsx"},
        ])

        eliminados = repo.eliminar_por_fuente("test.xlsx")
        assert eliminados == 2

        restantes = repo.buscar_todos()
        assert len(restantes) == 1


# ─────────────────────────────────────────
# Tests: DocumentIngester.ingestir_xlsx_precios()
# ─────────────────────────────────────────

class TestDocumentIngesterPrecios:
    """Tests de ingestión de precios desde Excel."""

    def test_ingestir_xlsx_precios_basico(self, db_session):
        """Ingesta básica de Excel de precios."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            _crear_excel_formato_simple(tmp.name)
            ruta = tmp.name

        cantidad = ingester.ingestir_xlsx_precios(ruta, "particular")

        assert cantidad > 0

        # Verificar que se crearon los precios
        precios = price_repo.buscar_todos(tipo_afiliacion="particular")
        assert len(precios) > 0

    def test_ingestir_xlsx_precios_formato_rango(self, db_session):
        """Ingesta de Excel con rangos de edad."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            _crear_excel_precios(tmp.name)
            ruta = tmp.name

        cantidad = ingester.ingestir_xlsx_precios(ruta, "particular")

        # Debería crear precios para cada plan/zona/rango
        assert cantidad > 0

        # Verificar precios específicos
        precio_medimax_cordoba = price_repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            edad=25,
        )
        assert precio_medimax_cordoba is not None
        assert precio_medimax_cordoba.precio == 18000.0

        precio_medimax_cordoba_35 = price_repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            edad=35,
        )
        assert precio_medimax_cordoba_35 is not None
        assert precio_medimax_cordoba_35.precio == 22000.0

    def test_ingestir_xlsx_multi_hojas(self, db_session):
        """Ingesta de Excel con múltiples hojas (3 tipos de afiliación)."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            _crear_excel_multi_hojas(tmp.name)
            ruta = tmp.name

        cantidad = ingester.ingestir_xlsx_precios(ruta, "particular")

        assert cantidad > 0

    def test_ingestir_xlsx_sin_price_repo(self, db_session):
        """Error si no se provee PriceRepository."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        ingester = DocumentIngester(engine)  # Sin price_repository

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            _crear_excel_formato_simple(tmp.name)
            ruta = tmp.name

        with pytest.raises(ValueError, match="PriceRepository"):
            ingester.ingestir_xlsx_precios(ruta, "particular")

    def test_ingestir_xlsx_archivo_no_existe(self, db_session):
        """Error si el archivo no existe."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        with pytest.raises(FileNotFoundError):
            ingester.ingestir_xlsx_precios("/no/existe.xlsx", "particular")

    def test_ingestir_xlsx_precios_parseo_montos(self, db_session):
        """Verificar parseo de montos con formato argentino."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        # Crear Excel con formato argentino
        wb = Workbook()
        ws = wb.active
        ws.title = "Particulares"
        ws.append(["Plan", "Córdoba"])
        ws.append(["Medimax", "$18.000"])
        ws.append(["Gold", "$35.000,50"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            wb.close()
            ruta = tmp.name

        cantidad = ingester.ingestir_xlsx_precios(ruta, "particular")
        assert cantidad == 2

        precio = price_repo.buscar_precio(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
        )
        assert precio is not None
        assert precio.precio == 18000.0

    def test_normalizar_nombre_plan(self, db_session):
        """Normalización de nombres de plan."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine

        engine = KnowledgeEngine(db_session)
        ingester = DocumentIngester(engine)

        assert ingester._normalizar_nombre_plan("Medimax CO") == "medimax_co"
        assert ingester._normalizar_nombre_plan("Medimax") == "medimax"
        assert ingester._normalizar_nombre_plan("Medimax Gold") == "medimax_gold"
        assert ingester._normalizar_nombre_plan("Gold") == "gold"
        assert ingester._normalizar_nombre_plan("Plan Joven") == "plan_joven"
        assert ingester._normalizar_nombre_plan("Joven") == "plan_joven"


# ─────────────────────────────────────────
# Tests: ServiredCalculator con PriceRepository
# ─────────────────────────────────────────

class TestServiredCalculatorPrecios:
    """Tests de la calculadora con precios estructurados."""

    def test_obtener_precio_tabla(self, db_session):
        """Obtener precio desde la tabla estructurada."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        precio = calc._obtener_precio_tabla(
            tipo_afiliacion="particular",
            nombre_plan="medimax",
            zona="cordoba",
        )
        assert precio == 18000.0

    def test_obtener_precio_tabla_con_edad(self, db_session):
        """Obtener precio con restricción de edad."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="plan_joven",
            zona="cordoba",
            precio=8000.0,
            edad_desde=18,
            edad_hasta=30,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)

        # Edad dentro del rango
        precio = calc._obtener_precio_tabla(
            tipo_afiliacion="particular",
            nombre_plan="plan_joven",
            zona="cordoba",
            edad=25,
        )
        assert precio == 8000.0

        # Edad fuera del rango
        precio = calc._obtener_precio_tabla(
            tipo_afiliacion="particular",
            nombre_plan="plan_joven",
            zona="cordoba",
            edad=35,
        )
        assert precio is None

    def test_calcular_valor_plan_con_tabla(self, db_session):
        """Calcular valor del plan usando precios de la tabla."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        valor = calc.calcular_valor_plan(
            nombre_plan="medimax",
            zona="cordoba",
            edades=[35],
            tipo_afiliacion="particular",
        )
        assert valor == 18000.0

    def test_calcular_valor_plan_integrantes_multiples(self, db_session):
        """Calcular valor del plan para múltiples integrantes."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        valor = calc.calcular_valor_plan(
            nombre_plan="medimax",
            zona="cordoba",
            edades=[35, 33, 8],
            tipo_afiliacion="particular",
        )
        assert valor == 54000.0

    def test_cotizar_con_precios_tabla(self, db_session):
        """Cotización completa usando precios de la tabla."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        lead = _crear_lead_mock(edad=35, tipo_afiliacion=TipoAfiliacion.PARTICULAR)

        resultado = calc.cotizar(
            lead=lead,
            zona="cordoba",
            nombre_plan="medimax",
        )

        assert resultado.valor_plan_total == 18000.0
        assert resultado.valor_a_pagar == 18000.0  # Sin aportes
        assert len(resultado.desglose_por_integrante) == 1
        assert resultado.desglose_por_integrante[0]["valor"] == 18000.0

    def test_cotizar_con_aportes_y_precios_tabla(self, db_session):
        """Cotización con aportes y precios de la tabla."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        lead = _crear_lead_mock(edad=35, tipo_afiliacion=TipoAfiliacion.PARTICULAR)

        resultado = calc.cotizar(
            lead=lead,
            conceptos_obra_social=[5000.0],
            zona="cordoba",
            nombre_plan="medimax",
        )

        # Aportes: 5000 * 33.33 * 7 / 100 = 11665.5
        assert resultado.aportes_calculados > 0
        assert resultado.valor_a_pagar < resultado.valor_plan_total

    def test_cotizar_fallback_a_knowledge(self, db_session):
        """Sin PriceRepository, usa KnowledgeRepository como fallback."""
        from app.services.servired_calculator import ServiredCalculator

        # Sin price_repository
        calc = ServiredCalculator(db_session)
        lead = _crear_lead_mock(edad=35)

        resultado = calc.cotizar(
            lead=lead,
            zona="cordoba",
            nombre_plan="medimax",
        )

        # Sin knowledge cargado, valor será 0
        assert resultado.valor_plan_total == 0.0

    def test_calcular_valor_plan_sin_tipo_afiliacion(self, db_session):
        """Sin tipo_afiliacion explícito, usa precio sin filtro de tipo."""
        from app.services.servired_calculator import ServiredCalculator

        price_repo = PriceRepository(db_session)
        price_repo.crear(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        valor = calc.calcular_valor_plan(
            nombre_plan="medimax",
            zona="cordoba",
            edades=[35],
        )
        # Sin tipo_afiliacion, cotizar usa 'particular' por default
        assert valor == 18000.0


# ─────────────────────────────────────────
# Tests: Integración completa
# ─────────────────────────────────────────

class TestIntegracionCompletaPrecios:
    """Tests de integración completa: Excel -> Repository -> Calculator."""

    def test_flujo_completo_excel_a_cotizacion(self, db_session):
        """
        Flujo completo:
        1. Crear Excel ficticio
        2. Ingestar precios
        3. Cotizar con calculator
        """
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine
        from app.services.servired_calculator import ServiredCalculator

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        # 1. Crear Excel ficticio
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            sheets = {
                "Particulares": [
                    ["Plan", "18-30 Córdoba", "18-30 Interior", "31+ Córdoba", "31+ Interior"],
                    ["Medimax CO", 12000, 10500, 15000, 13000],
                    ["Medimax", 18000, 16000, 22000, 19500],
                    ["Medimax Gold", 28000, 25000, 35000, 31000],
                    ["Gold", 35000, 31000, 42000, 37000],
                ]
            }
            _crear_excel_precios(tmp.name, "particular", sheets)
            ruta = tmp.name

        # 2. Ingestar precios
        cantidad = ingester.ingestir_xlsx_precios(ruta, "particular")
        assert cantidad > 0

        # 3. Cotizar
        calc = ServiredCalculator(db_session, price_repository=price_repo)
        lead = _crear_lead_mock(edad=25, tipo_afiliacion=TipoAfiliacion.PARTICULAR)

        resultado = calc.cotizar(
            lead=lead,
            zona="cordoba",
            nombre_plan="medimax",
        )

        assert resultado.valor_plan_total == 18000.0
        assert len(resultado.desglose_por_integrante) == 1
        assert resultado.desglose_por_integrante[0]["valor"] == 18000.0

    def test_flujo_multi_integrantes(self, db_session):
        """Flujo con múltiples integrantes y precios por edad."""
        from app.services.document_ingester import DocumentIngester
        from app.services.knowledge_engine import KnowledgeEngine
        from app.services.servired_calculator import ServiredCalculator

        engine = KnowledgeEngine(db_session)
        price_repo = PriceRepository(db_session)
        ingester = DocumentIngester(engine, price_repository=price_repo)

        # Crear Excel con rangos de edad
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            sheets = {
                "Particulares": [
                    ["Plan", "18-30 Córdoba", "31+ Córdoba"],
                    ["Medimax", 18000, 22000],
                ]
            }
            _crear_excel_precios(tmp.name, "particular", sheets)
            ruta = tmp.name

        # Ingestar
        ingester.ingestir_xlsx_precios(ruta, "particular")

        # Crear lead con familia
        lead = _crear_lead_mock(
            edad=25,
            tipo_afiliacion=TipoAfiliacion.PARTICULAR,
            conyuge=True,
        )

        calc = ServiredCalculator(db_session, price_repository=price_repo)
        resultado = calc.cotizar(
            lead=lead,
            zona="cordoba",
            nombre_plan="medimax",
        )

        # Titular (25 años) + Cónyuge
        assert resultado.valor_plan_total > 0
