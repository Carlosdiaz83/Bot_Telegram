"""
Tests Sprint 18A — Importador de precios SERVIRED.

Cubre:
    - PriceRepository.upsert(): creación, actualización, sin cambio
    - price_importer: lectura de Excel, detección de hojas, importación
    - ImportResult: estadísticas correctas
    - Idempotencia: re-ejecución sin duplicar
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database.models import Base, ServiredPriceDB
from app.database.repository import PriceRepository
from app.services.price_importer import (
    ImportResult,
    _buscar_columna_plan,
    _detectar_tipo_hoja,
    _mapear_columnas_precio,
    _normalizar_plan,
    _parsear_precio,
    importar_precios,
)


# ─────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────

@pytest.fixture()
def db_session():
    """Sesión de test con DB SQLite en memoria."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()
    engine.dispose()


@pytest.fixture()
def price_repo(db_session):
    """PriceRepository con sesión de test."""
    return PriceRepository(db_session)


# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def _crear_xlsx(
    ruta: str,
    sheets: dict | None = None,
) -> None:
    """Crea un archivo .xlsx ficticio de precios."""
    wb = Workbook()

    if sheets is None:
        sheets = {
            "Particulares": [
                ["Plan", "18-30 Córdoba", "18-30 Interior", "31-50 Córdoba", "31-50 Interior"],
                ["Medimax CO", 12000, 10500, 15000, 13000],
                ["Medimax", 18000, 16000, 22000, 19500],
            ],
        }

    first = True
    for nombre, filas in sheets.items():
        if first:
            ws = wb.active
            ws.title = nombre
            first = False
        else:
            ws = wb.create_sheet(nombre)
        for fila in filas:
            ws.append(fila)

    wb.save(ruta)
    wb.close()


def _crear_xls(
    ruta: str,
    sheets: dict | None = None,
) -> None:
    """Crea un archivo .xls ficticio de precios usando xlrd xlwt."""
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt no instalado, no se puede crear .xls de test")

    wb = xlwt.Workbook()

    if sheets is None:
        sheets = {
            "Particulares": [
                ["Plan", "18-30 Córdoba", "18-30 Interior", "31-50 Córdoba", "31-50 Interior"],
                ["Medimax CO", 12000, 10500, 15000, 13000],
                ["Medimax", 18000, 16000, 22000, 19500],
            ],
        }

    for nombre, filas in sheets.items():
        ws = wb.add_sheet(nombre)
        for i, fila in enumerate(filas):
            for j, val in enumerate(fila):
                ws.write(i, j, val)

    wb.save(ruta)


# ─────────────────────────────────────────
# Tests: _parsear_precio
# ─────────────────────────────────────────

class TestParsearPrecio:
    """Tests para _parsear_precio."""

    def test_entero(self):
        assert _parsear_precio(15000) == 15000.0

    def test_float(self):
        assert _parsear_precio(15.5) == 15.5

    def test_string_simple(self):
        assert _parsear_precio("15000") == 15000.0

    def test_string_con_punto_miles(self):
        assert _parsear_precio("15.000") == 15000.0

    def test_string_con_coma_decimal(self):
        assert _parsear_precio("15,5") == 15.5

    def test_string_formato_argentino(self):
        assert _parsear_precio("$15.000") == 15000.0

    def test_string_vacio(self):
        assert _parsear_precio("") == 0.0

    def test_none(self):
        assert _parsear_precio(None) == 0.0

    def test_string_con_espacios(self):
        assert _parsear_precio(" 15000 ") == 15000.0


# ─────────────────────────────────────────
# Tests: _normalizar_plan
# ─────────────────────────────────────────

class TestNormalizarPlan:
    """Tests para _normalizar_plan."""

    def test_medimax_co(self):
        assert _normalizar_plan("Medimax CO") == "medimax_co"

    def test_medimax(self):
        assert _normalizar_plan("Medimax") == "medimax"

    def test_medimax_gold(self):
        assert _normalizar_plan("Medimax Gold") == "medimax_gold"

    def test_gold(self):
        assert _normalizar_plan("Gold") == "gold"

    def test_minusculas(self):
        assert _normalizar_plan("medimax") == "medimax"


# ─────────────────────────────────────────
# Tests: _buscar_columna_plan
# ─────────────────────────────────────────

class TestBuscarColumnaPlan:
    """Tests para _buscar_columna_plan."""

    def test_header_con_plan(self):
        header = ["Plan", "18-30 Córdoba", "Interior"]
        assert _buscar_columna_plan(header) == 0

    def test_header_con_nombre(self):
        header = ["Nombre", "Precio"]
        assert _buscar_columna_plan(header) == 0

    def test_header_sin_plan(self):
        header = ["Zona", "Monto"]
        assert _buscar_columna_plan(header) == -1


# ─────────────────────────────────────────
# Tests: _mapear_columnas_precio
# ─────────────────────────────────────────

class TestMapearColumnasPrecio:
    """Tests para _mapear_columnas_precio."""

    def test_dos_zonas(self):
        header = ["Plan", "18-30 Córdoba", "18-30 Interior"]
        idx_plan = 0
        cols = _mapear_columnas_precio(header, idx_plan)
        assert len(cols) == 2
        assert cols[0]["zona"] == "cordoba"
        assert cols[1]["zona"] == "interior"

    def test_sin_zona(self):
        header = ["Plan", "Monto"]
        idx_plan = 0
        cols = _mapear_columnas_precio(header, idx_plan)
        assert len(cols) == 0


# ─────────────────────────────────────────
# Tests: _detectar_tipo_hoja
# ─────────────────────────────────────────

class TestDetectarTipoHoja:
    """Tests para _detectar_tipo_hoja."""

    def test_particulares(self):
        assert _detectar_tipo_hoja("PARTICULARES") == "particular"

    def test_monotributos(self):
        assert _detectar_tipo_hoja("MONOTRIBUTOS") == "monotributo"

    def test_relacion_dependencia(self):
        assert _detectar_tipo_hoja("RELACION DE DEPENDENCIA") == "relacion_dependencia"

    def test_desconocida(self):
        assert _detectar_tipo_hoja("OTRA COSA") is None


# ─────────────────────────────────────────
# Tests: PriceRepository.upsert
# ─────────────────────────────────────────

class TestPriceRepositoryUpsert:
    """Tests para PriceRepository.upsert."""

    def test_crear_nuevo(self, price_repo):
        accion, registro = price_repo.upsert(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=15000.0,
            fuente="test.xlsx",
        )
        assert accion == "created"
        assert registro.precio == 15000.0

    def test_actualizar_precio(self, price_repo):
        price_repo.upsert(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=15000.0,
            fuente="test.xlsx",
        )
        accion, registro = price_repo.upsert(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=18000.0,
            fuente="test_v2.xlsx",
        )
        assert accion == "updated"
        assert registro.precio == 18000.0
        assert registro.fuente == "test_v2.xlsx"

    def test_sin_cambio(self, price_repo):
        price_repo.upsert(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=15000.0,
            fuente="test.xlsx",
        )
        accion, registro = price_repo.upsert(
            tipo_afiliacion="particular",
            plan="medimax",
            zona="cordoba",
            precio=15000.0,
            fuente="test_v2.xlsx",
        )
        assert accion == "unchanged"
        assert registro.fuente == "test_v2.xlsx"


# ─────────────────────────────────────────
# Tests: importar_precios (.xlsx)
# ─────────────────────────────────────────

class TestImportarPreciosXlsx:
    """Tests para importar_precios con archivos .xlsx."""

    def test_importar_xlsx_basico(self, db_session):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "precios.xlsx"
            sheets = {
                "Particulares": [
                    ["Plan", "18-30 Córdoba", "18-30 Interior"],
                    ["Medimax CO", 12000, 10500],
                    ["Medimax", 18000, 16000],
                ],
            }
            _crear_xlsx(str(ruta), sheets)

            resultado = importar_precios(ruta, db_session)

            assert resultado.exitoso
            assert resultado.hojas_procesadas == 1
            assert resultado.precios_creados == 4

    def test_importar_xlsx_tres_hojas(self, db_session):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "precios.xlsx"
            sheets = {
                "Particulares": [
                    ["Plan", "18-30 Córdoba", "Interior"],
                    ["Medimax", 18000, 16000],
                ],
                "Monotributos": [
                    ["Plan", "Córdoba", "Interior"],
                    ["Medimax", 20000, 18000],
                ],
                "Relacion de Dependencia": [
                    ["Plan", "Córdoba", "Interior"],
                    ["Gold", 25000, 22000],
                ],
            }
            _crear_xlsx(str(ruta), sheets)

            resultado = importar_precios(ruta, db_session)

            assert resultado.exitoso
            assert resultado.hojas_procesadas == 3
            assert resultado.precios_creados == 6

    def test_importar_xlsx_idempotente(self, db_session):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "precios.xlsx"
            sheets = {
                "Particulares": [
                    ["Plan", "Córdoba"],
                    ["Medimax", 18000],
                ],
            }
            _crear_xlsx(str(ruta), sheets)

            resultado1 = importar_precios(ruta, db_session)
            assert resultado1.precios_creados == 1

            resultado2 = importar_precios(ruta, db_session)
            assert resultado2.precios_creados == 0
            assert resultado2.precios_sin_cambio == 1

    def test_importar_xlsx_actualiza_precio(self, db_session):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "precios.xlsx"
            sheets_v1 = {
                "Particulares": [
                    ["Plan", "Córdoba"],
                    ["Medimax", 18000],
                ],
            }
            _crear_xlsx(str(ruta), sheets_v1)
            importar_precios(ruta, db_session)

            sheets_v2 = {
                "Particulares": [
                    ["Plan", "Córdoba"],
                    ["Medimax", 20000],
                ],
            }
            _crear_xlsx(str(ruta), sheets_v2)
            resultado = importar_precios(ruta, db_session)

            assert resultado.precios_actualizados == 1
            repo = PriceRepository(db_session)
            precio = repo.buscar_precio("particular", "medimax", "cordoba")
            assert precio is not None
            assert precio.precio == 20000.0

    def test_archivo_no_existe(self, db_session):
        resultado = importar_precios("/no/existe.xlsx", db_session)
        assert not resultado.exitoso
        assert "no encontrado" in resultado.errores[0]


# ─────────────────────────────────────────
# Tests: ImportResult
# ─────────────────────────────────────────

class TestImportResult:
    """Tests para ImportResult."""

    def test_exitoso_sin_errores(self):
        r = ImportResult(archivo="test.xlsx")
        r.precios_creados = 5
        assert r.exitoso

    def test_con_errores(self):
        r = ImportResult(archivo="test.xlsx")
        r.errores.append("falló")
        assert not r.exitoso

    def test_resumen(self):
        r = ImportResult(archivo="test.xlsx")
        r.precios_creados = 3
        r.precios_actualizados = 1
        texto = r.resumen()
        assert "test.xlsx" in texto
        assert "3" in texto
